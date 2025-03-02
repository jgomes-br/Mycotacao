from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment
import datetime
import os

from django.conf import settings

from ..models import Projeto

from .tabela_cotacao import TabelaCotacao, EnumStatus

COLUNA_INICIAL = 4
LINHA_INICIAL = 6
COR_FINALIZADO = PatternFill(start_color="0000FF00", fill_type = "solid")
COR_SEM_ACORDO = PatternFill(start_color="00FF0000", fill_type = "solid")
COR_AGUARDANDO = PatternFill(start_color="00808080", fill_type = "solid")
modelo = os.path.join(settings.MEDIA_ROOT, "modelo.xlsx")

alinhamento_centro=Alignment(horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)

def ExportarExcel(endereco_arquivo, user, projeto: Projeto):
    wb = load_workbook(modelo)
    ws_cotacao = wb['Cotacao']
    ws_cotacao['B2'] = datetime.datetime.now()
    ws_cotacao['B3'] = projeto.nome

    linha = LINHA_INICIAL
    coluna = COLUNA_INICIAL

    for fornecedor in projeto.fornecedores.all():
        cell = ws_cotacao.cell(linha-1, coluna) 
        cell.value = fornecedor.nome_fornecedor
        cell.alignment = alinhamento_centro
        coluna += 1
    
    coluna = COLUNA_INICIAL
    tb = TabelaCotacao(user, projeto)
    for produto in tb.cotacoes:
        ws_cotacao.cell(linha, COLUNA_INICIAL-3).value = produto.descricao

        for cotacaco in tb.cotacoes[produto]:
            custo = cotacaco.preco
            custo += cotacaco.preco * (cotacaco.fornecedor.contrato/100)
            cell = ws_cotacao.cell(linha, coluna)
            cell.value = custo
            if cotacaco.status == EnumStatus.FINALIZADO_SUCESSO:
                cell.fill = COR_FINALIZADO
            elif cotacaco.status == EnumStatus.FINALIZDO_SEM_SUCESSO:
                cell.fill = COR_SEM_ACORDO
            elif cotacaco.status == EnumStatus.AGUARDANDO:
                cell.fill = COR_AGUARDANDO

            cell.number_format = "0.00"
            cell.alignment = alinhamento_centro
            coluna += 1

        coluna = COLUNA_INICIAL
        linha += 1

    wb.save(endereco_arquivo)